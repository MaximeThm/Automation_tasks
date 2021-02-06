import os
import openpyxl

root = '/Users/maximethomas/Desktop/Excel'
files = os.listdir(root)

for file in files:
    file = os.path.join(root, file)
    if file.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file)
        name = wb.sheetnames[0]
        name = ''.join(name)
        sheet = wb[name]
        status = sheet.cell(sheet.min_row, 1).value
        sheet.delete_rows(sheet.min_row, 1)
        wb.save(file)


